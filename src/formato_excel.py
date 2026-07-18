from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Crear libro de trabajo

def dar_formato(ws):
    # Encabezados
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="366092",
                                end_color="366092",
                                fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    # Ajustar ancho
    for columna in ws.columns:
        max_length = 0
        letra = get_column_letter(columna[0].column)

        for celda in columna:
            if celda.value:
                max_length = max(max_length, len(str(celda.value)))
        
        ancho = min(max_length + 5, 70)
        ws.column_dimensions[letra].width = ancho

    # Dimension de columna ofertas
    ws.column_dimensions["F"].width = 10

    # crear columna estado
    ws["G1"] = "Estado"

    ws["G1"].font = Font(bold=True, color="FFFFFF")
    ws["G1"].fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid")
    ws["G1"].alignment = Alignment(horizontal="center")
        

    # Color verde para filas revisadas
    verde = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    regla = FormulaRule(
        formula=['$G2="x"'],
        fill=verde
    )

    # Aplicar desde la fila 2 hasta la última solo si hay datos

    if ws.max_row > 1:
        ws.conditional_formatting.add(
            f"A2:F{ws.max_row}",
            regla
        )
    
    # Formato moneda para la columna presupuesto

    if ws.max_row > 1:
        for fila in range(2, ws.max_row + 1):
            ws[f"D{fila}"].number_format = '$#,##0'

    # Hipervinculos
    # iteramos la columna que tiene los link desde la fila 2 hasta la ultima (esto ya que la 1 son los encabezados) solo si hay valores
    if ws.max_row > 1:
        for fila in range(2, ws.max_row + 1):
                
            codigo = ws[f"A{fila}"].value
            celda = ws[f"F{fila}"]

            celda.hyperlink = f"https://buscador.mercadopublico.cl/ficha?code={codigo}"
            celda.value = "Revisar"
            celda.style = "Hyperlink"
